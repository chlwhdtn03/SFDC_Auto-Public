#!/usr/bin/env python
# coding: utf-8

# # 라이브러리 불러오기

# In[2]:


import tkinter
import tkinter.font
import tkinter.messagebox
from tkinter import filedialog
from tkinter import *
from tkinter.ttk import *

import selenium
from selenium import webdriver
from selenium.webdriver import ActionChains

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait

from selenium.common.exceptions import NoSuchElementException

import os
from datetime import datetime, timedelta
from win11toast import toast, notify

import xlwings as xw
from xlwings.constants import DeleteShiftDirection

import pandas as pd
import numpy as np

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder


# # 주문번호 생성 자동화

# In[3]:


def order_list(src):
    try:
        book = xw.Book(src)
        sheet = book.sheets[0]
        sheet.range('3:3').api.Delete(DeleteShiftDirection.xlShiftUp) 
        target = sheet.used_range.options(pd.DataFrame, index=False).value
        book.close()
        print("[정보] order.xlsx를 불러왔습니다.")
        return list(filter(None,target['URL']))
    except:
        tkinter.messagebox.showerror("파일 찾을 수 없음.", "%s 경로의 파일을 찾을 수 없습니다.\n혹은 잘못된 양식일 수 있습니다." % src)
        return None

def order_open_chrome(src, SET_AMOUNT = 5):
    if SET_AMOUNT == None:
        tkinter.messagebox.showerror("세트 당 개수 입력 필수", "세트 당 개수를 입력해주세요.")
        return;
    else:
        try:
            SET_AMOUNT = int(SET_AMOUNT)
        except:
            tkinter.messagebox.showerror("잘못된 형식", "슷자만 입력해주세요.")
            return
    urlist = order_list(src)
    if urlist is None:
        return
    service = webdriver.ChromeService(executable_path='chromedriver/chromedriver.exe')
    driver = webdriver.Chrome(service=service)
    url = urlist[0] # 주문 조회페이지로 바로 접속
    
    driver.get(url)
    notify("로그인 필요", "새로 열린 크롬창에 로그인이 필요합니다.\n로그인 후 프로그램에서 확인 버튼을 눌러주세요.")
    tkinter.messagebox.showinfo("로그인 필요", "새로 열린 크롬창에 로그인을 해주시고 확인 버튼을 눌러주세요.")

    driver.get(url)
    notify("공지 팝업 제거 필요", "공지 등의 팝업창이 띄워져있으면 반드시 '7일 이내 다시 보지 않기'를 눌러주세요.\n완료 후 프로그램에서 확인 버튼을 눌러주세요.")
    tkinter.messagebox.showinfo("공지 팝업 무시 필요", "만약 공지 등의 팝업창이 띄워져있으면 반드시 '7일 이내 다시 보지 않기'를 눌러주세요.") 

    process(urlist, SET_AMOUNT, driver, src)

def process(urlist, SET_AMOUNT, driver, src):
    wait = WebDriverWait(driver, 60)
    error = 0
    current_url = ''
    current_pos = 0
    selected = ''

    error = 0
    selected = ''
    current_pos = 0
    
    for url in urlist:
        if error:
            chkERR(error, current_url, current_pos, selected, SET_AMOUNT, driver, src)
            return
        current_pos = 0 # url 이동시 마다 초기화 필수
        current_url = url
        driver.get(url)
        
        # 혹시 이미 뭔가 기입되어있는 상태라면 정지 후 알림
        try: 
            for i in range(100):
                selected = driver.find_element(By.ID, 'j_id0:mainFrm:multipleOrder:outMultipleOrderList:%d:j_id86' % i).text
                if selected == 'ERROR':
                    error = 2
                    chkERR(error, current_url, current_pos, selected, SET_AMOUNT, driver, src)
                    return
                elif selected != '':
                    error = 4
                    chkERR(error, current_url, current_pos, selected, SET_AMOUNT, driver, src)
                    return
        except NoSuchElementException:
            pass
                
        while True:
            try: 
                selected = driver.find_element(By.ID, 'j_id0:mainFrm:multipleOrder:outMultipleOrderList:%d:j_id86' % current_pos).text
            except NoSuchElementException:
                if current_pos > 0:
                    break
                elif current_pos == 0:
                    error = 3
                    break
                
            if selected == 'ERROR':
                error = 2
                break
            for i in range(SET_AMOUNT):
                if selected != driver.find_element(By.ID, 'j_id0:mainFrm:multipleOrder:outMultipleOrderList:%d:j_id86' % (current_pos+i)).text:
                    error = 1
                    break
            if selected != '' and not error: # 계속해서 탐색 & 방금 직전에 주문저장해서 새로 발급된 주문번호에 문제가 없으면 POS 이동
                current_pos += SET_AMOUNT
            elif error: # 불일치 에러 발생
                chkERR(error, current_url, current_pos, selected, SET_AMOUNT, driver, src)
                return
            elif selected == '': 
                # 주문번호가 모두 빈칸인 곳까지 내려온 상태
                # 체크박스 선택 후, '주문저장' 클릭
                driver.find_element(By.NAME, 'j_id0:mainFrm:multipleOrder:outMultipleOrderList:%d:j_id81' % int(current_pos)).click()
                wait.until(lambda d : driver.find_element(By.ID, 'j_id0:mainFrm:multipleOrder:j_id65:j_id66:j_id67:submitBtn').click() or True)
                wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'blockPage'))) # 로딩창이 뜰때까지 대기 (로딩창 뜨는 속도보다 코드 실행되는 속도가 더 빨라서 멈춰줘야함)
                wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, 'blockPage'))) # 로딩창 닫히면 재개
    if not error:
        chkERR(error, current_url, current_pos, selected, SET_AMOUNT, driver, src)


def chkERR(error, current_url, current_pos, selected, SET_AMOUNT, driver, src):
    
    if error == 1:
        notify('작업을 중지합니다.',
              '주문번호가 일치하지 않는 항목이 있습니다.\n주문번호 : %s\n확인된 위치 : %d번째 세트항목\n프로그램에서 다시 시작할 수 있습니다.' % (selected, (current_pos/SET_AMOUNT)+1), 
              scenario='incomingCall')
        req = tkinter.messagebox.askyesno("주문 중지", "주문번호가 일치하지 않는 항목이 있습니다.\n%d번째 세트항목의 주문번호가 %s(이)가 아닙니다.\n%s를 새로 불러와서 다시 시작할까요?\n(종료할 경우 새로 로그인해야 합니다.)" % ((current_pos/SET_AMOUNT)+1, selected, src)) 

        print("[에러] 작업을 중지합니다.")
        print("[에러] 주문번호가 일치하지 않는 항목이 있습니다.")
        print("[에러] %d번째 칸의 주문번호가 %s(이)가 아닙니다." % (current_pos+1, selected))
        if req:
            process(order_list(src), SET_AMOUNT, driver, src)
    elif error == 2:
        notify('작업을 중지합니다.', '주문번호가 Error인 항목입니다.\n확인된 위치 : %d번째 세트항목\n프로그램에서 다시 시작할 수 있습니다.' % (current_pos+1), 
              scenario='incomingCall')
        req = tkinter.messagebox.askyesno("주문 중지", "주문번호가 Error인 항목입니다.\n현재 URL : %s\n에러 위치 : %d번째 세트항목\n%s를 새로 불러와서 다시 시작할까요?\n(종료할 경우 새로 로그인해야 합니다.)" % (current_url, (current_pos/SET_AMOUNT)+1, src)) 

        
        print("[에러] 작업을 중지합니다.")
        print("[에러] 주문번호가 ERROR인 항목입니다.")
        print("[에러] 현재 URL :", current_url)
        print("[에러] 에러 위치 :", current_pos+1,"번째 세트항목")
        if req:
            process(order_list(src), SET_AMOUNT, driver, src)
    
    elif error == 3:
        notify('작업을 중지합니다.', '현재 URL에 아무런 내용이 없습니다.\n해당 URL: %s\n프로그램에서 다시 시작할 수 있습니다.' % (current_url), 
              scenario='incomingCall')
        req = tkinter.messagebox.askyesno("주문 중지", "현재 URL에 아무런 내용이 없습니다.\n현재 URL : %s\n%s를 새로 불러와서 다시 시작할까요?\n(종료할 경우 새로 로그인해야 합니다.)" % (current_url, src)) 

        print("[에러] 현재 URL의 아무런 내용이 없습니다.")
        print("[에러] order.xlsx의 URL을 다시 한번 확인해 주시고 다시 실행해주세요.")
        print("[에러] 현재 URL :", current_url)
        if req:
            process(order_list(src), SET_AMOUNT, driver, src)

    elif error == 4:
        notify('작업을 중지합니다.', '이미 주문이 완료된 페이지 같습니다.\n현재 URL : %s' % (current_url), 
              scenario='incomingCall')
        req = tkinter.messagebox.askyesno("주문 중지", "이미 주문이 완료된 페이지같습니다.\n현재 URL : %s\n%s를 새로 불러와서 다시 시작할까요?\n(종료할 경우 새로 로그인해야 합니다.)" % (current_url, src)) 

        print("[에러] 현재 URL의 아무런 내용이 없습니다.")
        print("[에러] order.xlsx의 URL을 다시 한번 확인해 주시고 다시 실행해주세요.")
        print("[에러] 현재 URL :", current_url)
        if req:
            process(order_list(src), SET_AMOUNT, driver, src)
    
    else:
        notify('주문 완료', '모든 작업이 완료되었습니다.')
        tkinter.messagebox.showinfo("주문 완료", "모든 작업이 완료되었습니다.\n확인을 누르면 해당 크롬창이 닫힙니다.") 
        print('[정보] 모든 작업이 완료되었습니다.')





# # 주문진척 다운로드 자동화

# In[4]:


def download_auto(모델번호, 판매처코드, rawdate):
    dates = []
    dates_str = ''
    i = 1
    for line in rawdate.splitlines():
        try:
            print(line)
            if line == "":
                continue
            dates.append([datetime.strptime(line.split("-")[0], '%y%m%d').strftime("%Y. %#m. %#d"), 
                          datetime.strptime(line.split("-")[1], '%y%m%d').strftime("%Y. %#m. %#d")])
            dates_str += "%d번째 다운로드 구간 : %s ~ %s\n" % (i, dates[-1][0], dates[-1][1])
            i+=1
        except:
            tkinter.messagebox.showerror("잘못된 날짜 형식", "날짜를 형식에 맞게 입력해주셔야 합니다!\n날짜 형식(YYMMDD-YYMMDD):240301-241124")
            return

    
    if 판매처코드 == '':
        tkinter.messagebox.showerror("판매처코드 입력 필수", "판매처코드가 비어있습니다.")
        return
    if 모델번호 == '':
        tkinter.messagebox.showerror("모델번호 입력 필수", "모델번호가 비어있습니다.")
        return
    if len(dates) < 1:
        tkinter.messagebox.showerror("기간 입력 필수", "기간 란에 다운로드 받을 날짜 구간을 입력해주셔야 합니다.\n날짜 형식(YYMMDD-YYMMDD):240301-241124")
        return
    ask = tkinter.messagebox.askyesno("다운로드 준비 완료", "판매처코드 : %s\n모델번호 : %s\n\n-- 구간 --\n%s\n-- -- --\n\n이대로 다운로드를 시작할까요?"
                                 % (판매처코드, 모델번호, dates_str))
    if not ask:
        return

    dir = './download'
    
    if not os.path.exists(dir):
        os.makedirs(dir)

    try:
        service = webdriver.ChromeService(executable_path='chromedriver/chromedriver.exe')
    
        options = webdriver.ChromeOptions()
        prefs = {'download.default_directory' : os.path.abspath(dir)}
        options.add_experimental_option('prefs', prefs)
        
        driver = webdriver.Chrome(service=service, options=options)
    except:
        notify("chromedriver.exe 파일 찾을 수 없음", "chromedriver 폴더 내부에 chromedriver.exe 파일이 없습니다.")
        return

    url = "https://sec-b2b--c.vf.force.com/apex/OrderProgressStatusList?sfdc.tabName=01r28000000ox8N" # 주문 조회페이지로 바로 접속
    driver.get(url)
    notify("로그인 필요", "새로 열린 크롬창에 로그인이 필요합니다.\n로그인 후 프로그램에서 확인 버튼을 눌러주세요.")
    tkinter.messagebox.showinfo("로그인 필요", "새로 열린 크롬창에 로그인을 해주시고 확인 버튼을 눌러주세요.")
    driver.get(url)
    notify("공지 팝업 제거 필요", "공지 등의 팝업창이 띄워져있으면 반드시 '7일 이내 다시 보지 않기'를 눌러주세요.\n완료 후 프로그램에서 확인 버튼을 눌러주세요.")
    tkinter.messagebox.showinfo("공지 팝업 무시 필요", "만약 공지 등의 팝업창이 띄워져있으면 반드시 '7일 이내 다시 보지 않기'를 눌러주세요.") 


    for date in dates:
        driver.switch_to.window(driver.window_handles[0])
        driver.find_element(By.ID, 'j_id0:pageBody:pageForm:j_id39:j_id53:j_id54:j_id57:RequiredField:soldtoinfo_lkwgt').click()
        
        # 판매처 조회칸에 값 입력 (가장 마지막에 열린창)
        driver.switch_to.window(driver.window_handles[-1])
        driver.switch_to.frame('searchFrame')
        driver.find_element(By.NAME, 'lksrch').send_keys(판매처코드)
        driver.find_element(By.ID, 'lkenhmdSEARCH_ALL').click()
        driver.find_element(By.NAME, 'lksrch').send_keys(Keys.ENTER)
        
        driver.switch_to.window(driver.window_handles[-1])
        driver.switch_to.frame('resultsFrame')
        
        driver.find_element(By.CLASS_NAME, 'dataCell').click()
        
        ###
        
        driver.switch_to.window(driver.window_handles[0])
        driver.find_element(By.ID, 'j_id0:pageBody:pageForm:j_id39:j_id53:j_id73:j_id75_lkwgt').click()
        
        # 물품 조회칸에 값 입력 (가장 마지막에 열린창)
        driver.switch_to.window(driver.window_handles[-1])
        driver.switch_to.frame('searchFrame')
        driver.find_element(By.NAME, 'lksrch').send_keys(모델번호)
        driver.find_element(By.NAME, 'lksrch').send_keys(Keys.ENTER)
        
        driver.switch_to.window(driver.window_handles[-1])
        driver.switch_to.frame('resultsFrame')
        
        driver.find_element(By.CLASS_NAME, 'dataCell').click()
        
        driver.switch_to.window(driver.window_handles[0])
        
        driver.find_elements(By.CLASS_NAME, 'rt')[0].clear()
        driver.find_elements(By.CLASS_NAME, 'rt')[0].send_keys(date[0])
        
        driver.find_elements(By.ID, 'j_id0:pageBody:pageForm:j_id39:j_id53:j_id90:j_id93:RequiredField:j_id104:j_id105:j_id108')[0].clear()
        driver.find_elements(By.ID, 'j_id0:pageBody:pageForm:j_id39:j_id53:j_id90:j_id93:RequiredField:j_id104:j_id105:j_id108')[0].send_keys(date[1])
        driver.find_elements(By.ID, 'j_id0:pageBody:pageForm:j_id39:j_id53:j_id90:j_id93:RequiredField:j_id104:j_id105:j_id108')[0].send_keys(Keys.ENTER)
    
        wait = WebDriverWait(driver, 60)
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'blockPage'))) # 로딩창이 뜰때까지 대기 (로딩창 뜨는 속도보다 코드 실행되는 속도가 더 빨라서 멈춰줘야함)
        wait = WebDriverWait(driver, 60)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, 'blockPage')))
        wait = WebDriverWait(driver, 60) # 원래 wait 쓸때마다 초기화 해줘야하는지 모르겠음. 근데 없으면 무시하고 넘어가서 추가해둠
        wait.until(lambda d : driver.find_element(By.ID, 'j_id0:pageBody:pageForm:j_id39:j_id141:j_id142:xyz').click() or True)
        
    driver.switch_to.window(driver.window_handles[0])
    driver.find_elements(By.ID, 'j_id0:pageBody:pageForm:j_id39:j_id141:j_id142:xyz')[0].click()

    notify('다운로드 완료 직전', '다운로드가 거의 완료된 것으로 보입니다.\ndownload 폴더를 확인해주세요. 있어야 할 파일 개수는 %d개 입니다.' % (len(dates)), 
              scenario='incomingCall')
    print(driver.get_downloadable_files())


# # 다운로드 받은 주문 진척파일 **합본 자동화**

# In[5]:


def merge_auto():
    dir = './merge'
    result_arr = []

    if not os.path.exists(dir):
        os.makedirs(dir)

    files = os.listdir(dir)
    files.sort(key = lambda x : os.path.getmtime(dir + "/" + x))

    if len(files) == 0:
        tkinter.messagebox.showerror("합칠 파일 없음", "merge 폴더에 아무런 xlsx 파일이 없습니다.\n파일을 넣고 다시 시도해주세요.")
        return
    try:
        for file in files:
            book = xw.Book(dir + "/" + file)
            sheet = book.sheets[0]
            sheet.range('1:1').api.Delete(DeleteShiftDirection.xlShiftUp) 
            target = sheet.used_range.options(pd.DataFrame, index=False).value
            book.close()
            
            # 2줄로 된 제목을 위해 제목 변경 ( 하드코딩 수정 예정 )
            target.columns.values[9] = '인수거부'
            target.columns.values[10] = '불량'
            target.drop([0], axis=0, inplace=True)
        
            result_arr.append(target)
        
        output_target = target.head(0)
        output_target = pd.concat(result_arr)
        output_target.to_excel("합본.xlsx", engine="openpyxl", index=False)
        notify('합본 완료', '합본 생성이 완료되었습니다.')
        tkinter.messagebox.showinfo("합본 완료", "합본 생성이 완료되었습니다.")
    except:
        notify('합본 중 오류 발생', '파일 확장자가 xlsx가 아니거나 파일이 깨진 것 같습니다.\n파일을 확인 후 다시 시도해주세요.')
        tkinter.messagebox.showerror("합본 중 오류 발생", "파일 확장자가 xlsx가 아니거나 파일이 깨진 것 같습니다.\n파일을 확인 후 다시 시도해주세요.")
        return


# # 유효주문 분류 자동화

# In[9]:


def style_sep(v):
    return 'background-color:lightgreen;'

def seperate_auto(src):
    try:
        book = xw.Book(src)
        sheet = book.sheets[0]
        target = sheet.used_range.options(pd.DataFrame, index=False).value
        book.close()
        target['고객명'].head(5) # 없는 파일이거나 잘못된 양식일 수 있는 except 검출을 위한 코드
    except:
        tkinter.messagebox.showerror("파일 찾을 수 없음.", "%s 경로의 파일을 찾을 수 없습니다.\n혹은 잘못된 양식일 수 있습니다." % src)
        return
    notify("불러오기 성공", "파일을 불러왔습니다.\n작업이 완료되면 여기로 알려드리겠습니다.")
    
    #고객명+생일로 열 추가
    target['고객명생일'] = target['고객명'].str.split("_").str[1]+target['고객명'].str.split("_").str[2]

    # None으로 Null값 통일
    target = target.replace({np.nan: None})
    person_list = target['고객명생일']

    # 고객명생일 형식 아닌거 지우기 (재단명, Null값은 여기서 삭제됨)
    person_list = list(filter(None,person_list))

    #중복제거
    person_list = list(dict.fromkeys(person_list))

    i = 0
    value = 0
    arr_df = []
    extra_df = [] # 중복주문 의심 시트
    
    for p in person_list:
        count_zza = 0
        count_zzb_zfm = 0
        value = 0
        plist = target.loc[target['고객명생일'].str.contains(p, na=False, regex=False)]['주문유형'].tolist()
        for i in plist:
            if i == 'YKKR-ZFM': #설치 후 고장
                value -= 1
                count_zzb_zfm += 1
            elif i == 'YKA1-ZZB': #단순변심
                value -= 1
                count_zzb_zfm += 1
            elif i == 'YKB2-ZZA': #설치계약
                value += 1
                count_zza += 1
            # 1일 경우 일단 설치계약 된걸로 이것만 가져가면 됨!
            # 0과 같거나 보다 작을경우 설치되지 않았으니 걸러야함!
            # 1보다 클 경우 중복주문 된거니 마지막에 주문된걸 살리고 상단에 있는 주문을 걸러야함!
        
        if value >= 1:
            tmp_df = target.loc[target['고객명생일'].str.contains(p, na=False, regex=False) & target['주문유형'].str.contains('YKB2-ZZA', na=False, regex=False),:]
            if count_zza - count_zzb_zfm > 1:
                e_df = tmp_df
                for t in range(count_zzb_zfm):
                    e_df = e_df.drop(e_df.index[0])
                extra_df.append(e_df) # 중복 의심에 ZZA 전부 추가
            
            for i in range(count_zza-1):
                print("[정보]", p, "님의 일부 데이터를 삭제합니다.")
                tmp_df = tmp_df.drop(tmp_df.index[0]) # 맨 위에것만 지우면 가장 마지막인 맨 밑에만 남을테니..
            arr_df.append(tmp_df)
    
    output_target = pd.concat(arr_df)
    output_target = output_target.sort_index()
    output_target = output_target.drop('고객명생일', axis=1)
    # output_target.to_excel("output.xlsx", sheet_name='분류 완료', engine="openpyxl", index=False)
    if len(extra_df): # 중복 의심이 없으면 패스해야함! 에러뜸! (0이면 false입니다!)
        extra_target = pd.concat(extra_df)
        extra_target = extra_target.drop('고객명생일', axis=1)
    # extra_target.to_excel("output.xlsx", sheet_name='중복주문 의심', engine="openpyxl", index=False)
     
    target = target.drop('고객명생일', axis=1)

    writer = pd.ExcelWriter("output.xlsx", engine = 'openpyxl')
    target.to_excel(writer, sheet_name = 'raw', index=False)
    output_target.to_excel(writer, sheet_name = '분류 완료', index=False)
    if not len(extra_df): # 중복의심 없으면 제목열만 추가 (0이면 false입니다!)
        extra_target = output_target.head(0)
    extra_target.to_excel(writer, sheet_name = '중복주문 의심', index=False)
    
    workbook = writer.book
    worksheet = writer.sheets['raw'] # 날짜 셀 서식 변경.. 하드코딩 말고 방법이 없는거같음
    for cell in worksheet['A']:
        cell.number_format = 'yyyy-mm-dd'
    for cell in worksheet['M']:
        cell.number_format = 'yyyy-mm-dd'
    for cell in worksheet['N']:
        cell.number_format = 'yyyy-mm-dd'

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells: # not check merge_cells
                continue
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    worksheet = writer.sheets['분류 완료']
    for cell in worksheet['A']:
        cell.number_format = 'yyyy-mm-dd'
    for cell in worksheet['M']:
        cell.number_format = 'yyyy-mm-dd'
    for cell in worksheet['N']:
        cell.number_format = 'yyyy-mm-dd'

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells: # not check merge_cells
                continue
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width


    worksheet = writer.sheets['중복주문 의심']
    for cell in worksheet['A']:
        cell.number_format = 'yyyy-mm-dd'
    for cell in worksheet['M']:
        cell.number_format = 'yyyy-mm-dd'
    for cell in worksheet['N']:
        cell.number_format = 'yyyy-mm-dd'

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells: # not check merge_cells
                continue
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    
    # extra_target.stylde.map(style_sep).to_excel(writer, sheet_name = '중복주문 의심', index=False)
    writer.close()
    
    
    notify("분류 완료", "분류가 완료되었습니다. output.xlsx를 확인해주세요.")


# # 매핑 자동화

# In[10]:


def mapping_auto(mapping_src, separated_src):
    if not mapping_src:
        tkinter.messagebox.showerror("에러 발생", "고유키가 있는 파일을 선택해주세요.")
        return
    if not separated_src:
        tkinter.messagebox.showerror("에러 발생", "분류된 파일을 선택해주세요.")
        return
    try:
        book = xw.Book(mapping_src)
        sheet = book.sheets[0]
        mapping_target = sheet.used_range.options(pd.DataFrame, index=False).value
        book.close()
        mapping_target['세대주'].head(5) # 없는 파일이거나 잘못된 양식일 수 있는 except 검출을 위한 코드
        mapping_target['생년월일'].head(5) # 없는 파일이거나 잘못된 양식일 수 있는 except 검출을 위한 코드
        mapping_target['기준키'].head(5) # 없는 파일이거나 잘못된 양식일 수 있는 except 검출을 위한 코드
        mapping_target = mapping_target[['세대주','생년월일','기준키']] # 필요한 항목 이외 전부 제거
        book = xw.Book(separated_src)
        sheet = book.sheets[1]
        seperated_target = sheet.used_range.options(pd.DataFrame, index=False).value
        seperated_target = seperated_target[['고객명','주문번호','요구납기일','확정일자','배달완료','물류센터']] # 필요한 항목 이외 전부 제거
        book.close()
        seperated_target['고객명'].head(5) # 없는 파일이거나 잘못된 양식일 수 있는 except 검출을 위한 코드
    except Exception as e: 
        tkinter.messagebox.showerror("에러 발생", e)
        return
    notify("불러오기 성공", "파일을 불러왔습니다.\n작업이 완료되면 여기로 알려드리겠습니다.")
    
    mapping_target['생년월일'] = mapping_target['생년월일'].astype(str)
    mapping_target['기준키'] = mapping_target['기준키'].astype(str)

    mapping_target['생년월일'] = mapping_target['생년월일'].str.split('.').str[0]
    mapping_target['기준키'] = mapping_target['기준키'].str.split('.').str[0]

    #고객명+생일로 열 추가
    mapping_target['고객명생일'] = mapping_target['세대주'].str[:]+mapping_target['생년월일'].str[-6:]

    seperated_target['고객명생일'] = (seperated_target['고객명'].str.split("_").str[1]+seperated_target['고객명'].str.split("_").str[2]).str.replace('）',")").str.replace('（',"(")

    result = mapping_target.merge(seperated_target, on='고객명생일', how='outer')
    result = result.drop('세대주', axis=1)
    result = result.drop('생년월일', axis=1)
    result = result.drop('고객명생일', axis=1)
    result = result[result['배달완료'] != 0]
    result = result[['기준키','고객명','주문번호','요구납기일','확정일자','배달완료']]
    
    result['기준키'] = pd.to_numeric(result['기준키'], errors='coerce').fillna(-1)

    result = result.sort_values(by='기준키', axis=0)
    result.loc[result['기준키'] == -1, '기준키']= '확인 필요'
    
    writer = pd.ExcelWriter("주문번호 및 기준키 매핑.xlsx", engine = 'openpyxl')
    result.to_excel(writer, sheet_name='매핑 결과', index=False)
    workbook = writer.book
    worksheet = writer.sheets['매핑 결과'] # 날짜 셀 서식 변경.. 하드코딩 말고 방법이 없는거같음

    
    for cell in worksheet['D']:
        cell.number_format = 'yyyy-mm-dd'
    for cell in worksheet['E']:
        cell.number_format = 'yyyy-mm-dd'

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells: # not check merge_cells
                continue
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    
    writer.close()
    
    notify("매핑 완료", "매핑이 완료되었습니다. '주문번호 및 기준키 매핑.xlsx'를 확인해주세요.")


# # 화면 생성 코드

# In[12]:


window = tkinter.Tk()
window.title("SFDC 자동화 (Build. 2024.10.15)")
window.resizable(False, False)

font = tkinter.font.Font(family="맑은 고딕", size=11)

s = tkinter.ttk.Style()
s.configure('TNotebook.Tab', font=('맑은 고딕','12'))
s.configure('TButton', font =('맑은 고딕', 12), padding=5)
s.configure('TLabel', font =('맑은 고딕', 10))
notebook = Notebook(window)
notebook.pack(fill=tkinter.BOTH)

# download_frame = tkinter.LabelFrame(window, text = "다운로드 자동화", padx=5)
# merge_frame = tkinter.LabelFrame(window, text = "합본 자동화", padx=5)
# seperate_frame = tkinter.LabelFrame(window, text = "분류 자동화", padx=5)
# order_frame = tkinter.LabelFrame(window, text = "주문 자동화", padx=5)
download_frame = tkinter.Frame(window, padx=10, pady=10)
merge_frame = tkinter.Frame(window, padx=10, pady=10)
seperate_frame = tkinter.Frame(window, padx=10, pady=10)
order_frame = tkinter.Frame(window, padx=10, pady=10)
mapping_frame = tkinter.Frame(window, padx=10, pady=10)

# CONSOLE <-- 활용공간 없음. 추가 X
# output_frame = tkinter.LabelFrame(window, text = "Console")
# output_frame.pack(side=tkinter.BOTTOM, fill=tkinter.BOTH)
# console_text = tkinter.Text(output_frame)
# console_text.pack(fill=tkinter.BOTH)
# console_text.bind("<Key>", lambda e: "break")

# # 배치 순서 조정
# order_frame.pack(side=tkinter.LEFT, fill=tkinter.BOTH, padx=5, pady=2)
# download_frame.pack(side=tkinter.LEFT, fill=tkinter.BOTH, padx=5, pady=2)
# merge_frame.pack(side=tkinter.LEFT, fill=tkinter.BOTH, padx=5, pady=2)
# seperate_frame.pack(side=tkinter.LEFT, fill=tkinter.BOTH, padx=5, pady=2)
# ###

notebook.add(order_frame, text='주문번호 생성')
notebook.add(download_frame, text='주문진척 다운로드')
notebook.add(merge_frame, text='합본 생성')
notebook.add(seperate_frame, text='주문 분류')
notebook.add(mapping_frame, text='주문번호 및 기준키 매핑')




# download(주문진척 다운로드) 입력창 
download_field_sellerCode_frame = tkinter.Frame(download_frame, pady=2)
download_field_sellerCode_label = tkinter.Label(download_field_sellerCode_frame, width=10, text="판매처코드", font=font)
download_field_sellerCode = Entry(download_field_sellerCode_frame, font=font)

download_field_sellerCode_frame.pack()
download_field_sellerCode_label.pack(side=tkinter.LEFT)
download_field_sellerCode.pack(fill=tkinter.BOTH)
                                                
download_field_modelNum_frame = tkinter.Frame(download_frame, pady=2)
download_field_modelNum_label = tkinter.Label(download_field_modelNum_frame, width=10, text="모델번호", font=font)
download_field_modelNum = Entry(download_field_modelNum_frame, font=font)

download_field_modelNum_frame.pack()
download_field_modelNum_label.pack(side=tkinter.LEFT)
download_field_modelNum.pack(fill=tkinter.BOTH)

download_field_date_frame = tkinter.Frame(download_frame, pady=2)
download_field_date_label = tkinter.Label(download_field_date_frame, width=10, text="기간", font=font)
scrollbar = Scrollbar(download_field_date_frame)

download_field_date = Text(download_field_date_frame, font=font, width=20, height=10, yscrollcommand = scrollbar.set)

download_field_date_frame.pack()
download_field_date_label.pack(side=tkinter.LEFT)
download_field_date.pack(side=tkinter.LEFT)
scrollbar.pack(fill=tkinter.BOTH, side=tkinter.RIGHT)
scrollbar["command"]=download_field_date.yview
# download_field_date.pack(fill=tkinter.BOTH)



# order(주문자동화) 파일 선택기
order_file_label = tkinter.Label(order_frame, width=20, height=3, wraplength=350, justify='center', text="파일을 선택해주세요", font=font)
order_file_btn = Button(order_frame, command=lambda: order_file_label.config(text=filedialog.askopenfilename(filetypes=(("엑셀(xlsx) 파일", "*.xlsx"),))), text="주문 파일 선택")

order_file_btn.pack()
order_file_label.pack(fill=tkinter.BOTH)
Separator(order_frame, orient="horizontal").pack(fill=tkinter.BOTH, pady=10)


#seperate file
seperate_file_label = tkinter.Label(seperate_frame, width=20, height=3, wraplength=350, justify='center', text="파일을 선택해주세요", font=font)
seperate_file_btn = Button(seperate_frame, command=lambda: seperate_file_label.config(text=filedialog.askopenfilename(filetypes=(("엑셀(xlsx) 파일", "*.xlsx"),))), text="파일 선택")
seperate_file_btn.pack()
seperate_file_label.pack(fill=tkinter.BOTH)
Separator(seperate_frame, orient="horizontal").pack(fill=tkinter.BOTH, pady=10)


#mapping file
mapping_key_file_label = tkinter.Label(mapping_frame, width=20, height=3, wraplength=350, justify='center', text="고유키가 있는 파일을 선택해주세요", font=font)
mapping_key_file_btn = Button(mapping_frame, command=lambda: mapping_key_file_label.config(text=filedialog.askopenfilename(filetypes=(("엑셀(xls) 파일", "*.xls"),))), text="고유키 파일 선택")
mapping_key_file_btn.pack()
mapping_key_file_label.pack(fill=tkinter.BOTH)
Separator(mapping_frame, orient="horizontal").pack(fill=tkinter.BOTH, pady=10)
mapping_separated_file_label = tkinter.Label(mapping_frame, width=20, height=3, wraplength=350, justify='center', text="분류 완료된 파일을 선택해주세요", font=font)
mapping_separated_file_btn = Button(mapping_frame, command=lambda: mapping_separated_file_label.config(text=filedialog.askopenfilename(filetypes=(("엑셀(xlsx) 파일", "*.xlsx"),))), text="분류된 파일 선택")
mapping_separated_file_btn.pack()
mapping_separated_file_label.pack(fill=tkinter.BOTH)
Separator(mapping_frame, orient="horizontal").pack(fill=tkinter.BOTH, pady=10)





download_auto_button = Button(download_frame, command=lambda: download_auto(download_field_modelNum.get().strip(), download_field_sellerCode.get().strip(), download_field_date.get("1.0",'end-1c').strip()), width=30, text="다운로드 자동화 시작")
download_auto_button.pack(pady=3) 

merge_auto_button = Button(merge_frame, command=lambda: merge_auto(), width=30, text="합본 자동화 시작")
merge_auto_button.pack() 

seperate_auto_button = Button(seperate_frame, command=lambda: seperate_auto(seperate_file_label.cget("text")), width=30, text="분류 자동화 시작")
seperate_auto_button.pack()

order_auto_button = Button(order_frame, command=lambda: order_open_chrome(order_file_label.cget("text")), width=30, text="주문(세트 당 5개) 자동화 시작")
order_auto_button.pack() 

mapping_auto_button = Button(mapping_frame, command=lambda: mapping_auto(mapping_key_file_label.cget("text"), mapping_separated_file_label.cget("text")), width=30, text="고유키 매핑 시작")
mapping_auto_button.pack() 









# order(주문자동화) 입력창


Separator(order_frame, orient="horizontal").pack(fill=tkinter.BOTH, pady=10)
order_field_setNum_frame = tkinter.Frame(order_frame, pady=5)
order_field_setNum_label = tkinter.Label(order_field_setNum_frame, width=10, text="세트 당 개수", font=font)
order_field_setNum = Entry(order_field_setNum_frame, font=font)

order_field_setNum_label.pack(side=tkinter.LEFT)
order_field_setNum.pack(fill=tkinter.BOTH)
order_field_setNum_frame.pack()

order_auto_customized_button = Button(order_frame, command=lambda: order_open_chrome(order_file_label.cget("text"), order_field_setNum.get()), width=30, text="주문(개수 입력형) 자동화 시작")
order_auto_customized_button.pack() 

download_description = Label(download_frame, text="download 폴더에 저장됩니다.")
merge_description = Label(merge_frame, text="merge 폴더에 같은 양식의 xlsx 파일들을 넣어두면\n합본.xlsx 파일이 생성됩니다.")
seperate_description = Label(seperate_frame, text="선택한 엑셀 파일에서 유효한 ZZA 주문만 추출하여\noutput.xlsx 파일을 생성합니다.")
order_description = Label(order_frame, text="선택한 엑셀 파일의 URL에 접속하여 주문번호를 생성합니다.")
mapping_description = Label(mapping_frame, text="고유키가 들어있는 엑셀 파일과 분류가 완료된 엑셀파일을\n선택하면 서로 매핑된 엑셀 파일이 생성됩니다.")

download_description.pack(side=tkinter.BOTTOM)
merge_description.pack(side=tkinter.BOTTOM)
seperate_description.pack(side=tkinter.BOTTOM)
order_description.pack(side=tkinter.BOTTOM)
mapping_description.pack(side=tkinter.BOTTOM)




window.mainloop()
## GUI 로드 끝


# In[ ]:




